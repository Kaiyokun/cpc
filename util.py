#!/usr/bin/python
# -*- coding: utf8 -*-


# 参考: https://automatetheboringstuff.com/chapter10/
import logging as log

# 配置用于调试的log对象
log.basicConfig(level=log.DEBUG, format=' %(asctime)s - %(levelname)s - '
	'%(thread)d - %(module)s - %(funcName)s - %(lineno)d - %(message)s')

# 从json中获取utf8编码的字符串而不是默认的unicode
# 参考: https://stackoverflow.com/questions/956867/how-to-get-string-objects-instead-of-unicode-from-json
import json

def json_load_byteified(file_handle):
	return _byteify(
		json.load(file_handle, object_hook=_byteify),
		ignore_dicts=True
	)

def json_loads_byteified(json_text):
	return _byteify(
		json.loads(json_text, object_hook=_byteify),
		ignore_dicts=True
	)

def _byteify(data, ignore_dicts = False):
	# if this is a unicode string, return its string representation
	if isinstance(data, unicode):
		return data.encode('utf-8')
	# if this is a list of values, return list of byteified values
	if isinstance(data, list):
		return [ _byteify(item, ignore_dicts=True) for item in data ]
	# if this is a dictionary, return dictionary of byteified keys and values
	# but only if we haven't already byteified it
	if isinstance(data, dict) and not ignore_dicts:
		return {
			_byteify(key, ignore_dicts=True): _byteify(value, ignore_dicts=True)
			for key, value in data.iteritems()
		}
	# if it's anything else, return it in its original form
	return data


# 以下是我的代码
import docx             # https://python-docx.readthedocs.io/en/latest/
import openpyxl as xlsx # http://openpyxl.readthedocs.io/en/default/

from bs4         import BeautifulSoup # https://www.crummy.com/software/BeautifulSoup/bs4/doc.zh/
from bs4.element import Tag

class Util:

	@staticmethod
	def read_conf_from_json(file_name):

		with open(file_name, 'rb') as file:

			return json_load_byteified(file)

	@staticmethod
	def write_conf_to_json(config, file_name):

		with open(file_name, 'wb') as file:

			# ensure_ascii - 防止把UTF-8转换为转义的Unicode格式(\u????)
			# indent       - 保持json的可读格式
			json.dump(config, file, ensure_ascii=False, indent=2)

	@staticmethod
	def read_conf_from_excel(file_name, sheet_name=None, ref=(1,1)):

		# 以只读模式打开excel
		wb = xlsx.load_workbook(file_name, read_only=True)

		# 未指定sheet名则打开默认sheet
		ws = wb.active if sheet_name is None else wb[sheet_name]

		# 将excel表格中除表头以外的每一行数据转换为key-value pairs

		# 默认表头为key set
		keys = [key.value.encode('UTF-8') for key in list(ws.get_squared_range(
			ref[1], ref[0], ws.max_column, ref[0]))[0]]

		# 表头以外的所有行作为一个json array
		json_array = []
		for cells in ws.get_squared_range(
			ref[1], ref[0] + 1, ws.max_column, ws.max_row):

			# 表头以外的每一行作为一个json object
			json_object = {}
			for i in range(len(cells)):

				# 将value转换为字符串并进行UTF-8编码
				json_object[keys[i]] = unicode(cells[i].value).encode('UTF-8')

			json_array += [json_object]

		wb.close()

		return json_array

	@staticmethod
	def create_html_from_template(file_name, args):

		with open(file_name, 'rb') as template:

			dom = BeautifulSoup(template, 'html.parser')

		# html模板中的占位符使用标签<why id='xxx'></why>包裹:)
		for tag in dom.select('why'):

			# 以标签<why id='xxx'></why>的id属性值xxx作为args的key
			tag.string = str(args[tag['id']])

		return str(dom)

	@staticmethod
	def create_html_list(file_name, list_data, list_text=''):

		with open(file_name, 'rb') as template:

			dom = BeautifulSoup(template, 'html.parser')

		# <table id="list"></table>
		table = dom.find('table', id='list')

		# <table id="list">
		#	<pre id="title">list_text</pre>
		# </table>
		pre = dom.new_tag('pre', id='title')
		pre.string = list_text
		table.append(pre)

		# <table id="list">
		#	<pre id="title">list_text</pre>
		#	<th><!-- list_data[0] -->
		#		<td>str(list_data[0][0])</td>
		#		<td>str(list_data[0][1])</td>
		#		...
		#	</th>
		#	<tr><!-- list_data[1] -->
		#		<td>str(list_data[1][0])</td>
		#		<td>str(list_data[1][1])</td>
		#		...
		#	</tr>
		#	<tr><!-- list_data[2] -->
		#		<td>str(list_data[2][0])</td>
		#		<td>str(list_data[2][1])</td>
		#		...
		#	</tr>
		#	...
		# </table>
		tr = dom.new_tag('tr')
		for col in list_data[0]:

			th        = dom.new_tag('th')
			th.string = str(col)

			tr.append(th)

		table.append(tr)

		for row in list_data[1:]:

			tr = dom.new_tag('tr')
			for col in row:

				td        = dom.new_tag('td')
				td.string = str(col)

				tr.append(td)

			table.append(tr)

		return str(dom)

	@staticmethod
	def create_html_exer_list(file_name, list_data, list_text=''):

		with open(file_name, 'rb') as template:

			dom = BeautifulSoup(template, 'html.parser')

		# <table id="list"></table>
		table = dom.find('table', id='list')

		# <table id="list">
		#	<pre id="title">list_text</pre>
		# </table>
		pre = dom.new_tag('pre', id='title')
		pre.string = list_text
		table.append(pre)

		# <table id="list">
		#	<pre id="title">list_text</pre>
		#	<th><!-- list_data[0] -->
		#		<td>str(list_data[0][0])</td>
		#		<td>str(list_data[0][1])</td>
		#		<td>str(list_data[0][2])</td>
		#	</th>
		#	<tr><!-- list_data[1] -->
		#		<td>str(list_data[1][0])
		#		<td><pre>str(list_data[1][1])</pre></td>
		#		<td><a href=str(list_data[1][2])>点击进入</a></td>
		#	</tr>
		#	<tr><!-- list_data[2] -->
		#		<td>str(list_data[2][0])
		#		<td><pre>str(list_data[2][1])</pre></td>
		#		<td><a href=str(list_data[2][2])>点击进入</a></td>
		#	</tr>
		#	...
		# </table>
		tr = dom.new_tag('tr')
		for col in list_data[0]:

			th        = dom.new_tag('th')
			th.string = str(col)

			tr.append(th)

		table.append(tr)

		for col in list_data[1:]:

			tr = dom.new_tag('tr')

			td        = dom.new_tag('td')
			td.string = str(col[0])
			tr.append(td)

			td         = dom.new_tag('td')
			pre        = dom.new_tag('pre')
			pre.string = str(col[1])
			td.append(pre)
			tr.append(td)

			td       = dom.new_tag('td')
			a        = dom.new_tag('a', href=str(col[2]))
			a.string = '点击进入'
			td.append(a)
			tr.append(td)

			table.append(tr)

		return str(dom)

	@staticmethod
	def create_html_test_detail(file_name, ex_id,
		test_detail_name='test_detail.xml'):

		# 创建新的html table row
		def new_row(content_list, is_table_header=False):

			tr  = dom.new_tag('tr')
			col = 'th' if is_table_header else 'td'

			for content in content_list:

				tag = dom.new_tag(col)

				# 指定了父标签属性
				if isinstance(content, list):

					for attr in content[1].keys():

						tag[attr] = content[1][attr]

					content = content[0]

				# 标签内容为html tag
				if isinstance(content, Tag):

					tag.append(content)

				# 标签内容为文本, 数值等
				else:

					tag.string = str(content)

				tr.append(tag)

			return tr

		with open(file_name, 'rb') as template:

			dom = BeautifulSoup(template, 'html.parser')

		# <why id="ex_id"></why>
		dom.find('why', id='ex_id').string = ex_id

		# <table id="test_detail"></table>
		table = dom.find('table', id='test_detail')

		# 表格标题行
		table.append(new_row(['测试集合', '状态', '测试用例', '状态',
			'测试用时(秒)', '错误信息'], is_table_header=True))

		with open(test_detail_name, 'rb') as test_detail:

			testsuites = BeautifulSoup(test_detail, 'xml')

		# 先提取跨行的testsuite信息
		for testsuite in testsuites.find_all('testsuite'):

			span_row_content = [
				[testsuite['name'], {'rowspan': testsuite['tests']}],
				['通过',            {'rowspan': testsuite['tests'],
					'bgcolor': 'green'}]]

			if testsuite.failure:

				span_row_content[1][0] = '失败'
				span_row_content[1][1]['bgcolor'] = 'red'

			# 跨行的tr只使用一次
			tr_once = True

			# 再逐行处理testcase信息
			for testcase in testsuite.find_all('testcase'):

				row_content = [
					testcase['name'],
					['通过', {'bgcolor': 'green'}],
					testcase['time'],
					'']

				if testcase.failure:

					row_content[1] = ['失败', {'bgcolor': 'red'}]
					row_content[3] = dom.new_tag('pre')
					row_content[3].string = '\n\n'.join([msg['message']
						for msg in testcase.find_all('failure')])

				if tr_once:

					tr_once = False

					# 添加首行的testsuite信息
					row_content = span_row_content + row_content

				table.append(new_row(row_content))

		return str(dom)

	@staticmethod
	def create_docx_table(file_name, args, output_file_name):

		with open(file_name, 'rb') as template:

			doc = docx.Document(template)

		for table in doc.tables:

			for row in table.rows:

				for cell in row.cells:

					# docx模板中的占位符使用#xxx
					if cell.text.startswith('#'):

						# 以#之后的xxx作为args的key
						cell.text = str(args[cell.text[1:]]).decode('UTF-8')

		doc.save(output_file_name)
